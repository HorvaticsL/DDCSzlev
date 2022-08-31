"""
SAP cikkeket tartalmazó, Excel táblázat adatainak 
felvezetése az SQL adatbázisba.

Készült: 2022.08.04

Utolsó módosítás dátuma: 2022.08.04
verzió: 01

"""

import sys
import ctypes
from types import NoneType
from openpyxl import load_workbook

import ini_fajl as inif
import datumido as di
import adatbconnect as dbconnect


def SAPCikkek_feltoltese(initomb, logfile):

    logfile.info('Adatbázis műveletek: SAPCikkek')
    prgneve = inif.initomb_eleme(initomb, 0)

    utdijmappa = inif.initomb_eleme(initomb, 11)
    utdijfile = inif.initomb_eleme(initomb, 12)
    sapcikk_sheet = inif.initomb_eleme(initomb, 48)
    
    odbc_driver = inif.initomb_eleme(initomb, 51)
    logfile.info('ODBC driver: %s', str(odbc_driver))
    sqlszerver = inif.initomb_eleme(initomb, 52)
    logfile.info('SQL szerver: %s', str(sqlszerver))
    sqladatb = inif.initomb_eleme(initomb, 53)
    logfile.info('Adatbázis: %s', str(sqladatb))
    sqlSAPCikkek = inif.initomb_eleme(initomb, 54)
    logfile.info('Adatbázis-tábla: %s', str(sqlSAPCikkek))

    utdijfajl = str(utdijmappa) + str(utdijfile)
    logfile.info('Excelfájl (SAP cikkek): %s', str(utdijfajl))

    FelvezetesDatuma = di.mainap("kotojel")

    try:

        logfile.info('Excel fájl megnyitása a SAP cikkekhez')
        utdij_book = load_workbook(filename=utdijfajl, data_only=True)
        sapcikksheet = utdij_book[sapcikk_sheet]
        
        logfile.info('ODBC kapcsolat megnyitása')
        kapcs =  dbconnect.ODBC_Kapcsolat(odbc_driver, sqlszerver, sqladatb)
        kapcscursor = kapcs.cursor()

        # Táblában lévő adatok törlése (DELETE ALL)
        logfile.info('SAPCikkek tábla adatainak törlése')
        kapcscursor.execute("DELETE FROM " + str(sqlSAPCikkek))
        kapcscursor.commit()

        maxsor = sapcikksheet.max_row
        logfile.info('Sorok száma (szállítólevelek): %s', str(maxsor))
        maxoszlop = sapcikksheet.max_column
        logfile.info('Oszlopok száma (szállítólevelek): %s', str(maxoszlop))

        for sor in range(2, maxsor+1):
            aktsor = str(sor)

            if type(sapcikksheet['A'+aktsor].value) == NoneType:
                break

            logfile.info("Cikkszám: %s", str(sapcikksheet['A'+aktsor].value))

            if type(sapcikksheet['I'+aktsor].value) == NoneType:
                megjegyzes = ''
            else:
                megjegyzes = sapcikksheet['I'+aktsor].value

            sql = (
                "INSERT INTO SAPCikkek VALUES( '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')"
                % (
                    sapcikksheet['A'+aktsor].value,
                    sapcikksheet['B'+aktsor].value,
                    sapcikksheet['C'+aktsor].value,
                    sapcikksheet['D'+aktsor].value,
                    sapcikksheet['E'+aktsor].value,
                    sapcikksheet['F'+aktsor].value,
                    sapcikksheet['G'+aktsor].value,
                    sapcikksheet['H'+aktsor].value,
                    megjegyzes,
                    FelvezetesDatuma,
                )
            )

            kapcscursor.execute(sql)
            kapcs.commit()

        # ***** for sor in range(2, maxsor+1): VÉGE

    except Exception as merror:
        logfile.error(
            'Ismeretlen hiba típusa, leírás: %s: %s', str(type(merror)), str(merror))
        logfile.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Ismeretlen hiba!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)

    # ODBC kapcsolat bezárása
    kapcs.close()
    logfile.info('SAPCikkek feltöltöltve, adatbáziskapcsolat bezárása.')
    # ***** SAPCikkek_feltoltese VÉGE
