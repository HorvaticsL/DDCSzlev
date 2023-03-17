"""
Excel fájl tartalmának feldolgozása
A fájlban lévő adatokat úgy kell átalakítani, hogy
az adatbázisba lementhetők legyenek - helységnevek, fuvar-, útdíjak, stb.
Készült: 2022.06.13

Visszaadja a függvény futása után a létrehozott, feldölgozott
Excel fájl nevét

Visszaadási érték: saveas_szlevfajl - útvonallal együtt

Utolsó módosítás dátuma: 2023.03.17
verzió: 04

"""
#import openpyxl
#from encodings.utf_8 import encode
#from imp import load_dynamic
import sys
import ctypes
from types import NoneType
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
#from openpyxl.utils import get_column_letter

import read_config as readcfg
import datumido as di
import make_pivottabla as mpt


def excelfajl_modositas(inifajl, logfile):
    # Utolsó módosítás dátuma: 2022.09.15

    logfile.info('Exel forrásfájl feldolgozása elindult')

    # INI fájladatok megnyitása
    config = readcfg.read_config(inifajl)

    # INI fájladatok beolvasása
    prgneve = config['EXE-File']['exe-file neve']
    # szállítólevelek forrás fájl
    excelfile = config['SAP-Excel']['sapexcel file neve']
    excelmappa = config['SAP-Excel']['sapexcel mappa']
    wbsheetneve = config['SAP-Excel']['sapexcel munkalap']

    # helységnevek
    helysegmappa = config['Helysegnev-EXCEL']['helysegnev mappa']
    helysegfile = config['Helysegnev-EXCEL']['helysegnev file neve']
    helysegsheet = config['Helysegnev-EXCEL']['helysegnev munkalap']
    helysegbalfelso = config['Helysegnev-EXCEL']['helysegnev tartomany bal-felso']
    helysegjobbalso = config['Helysegnev-EXCEL']['helysegnev tartomany jobb-also']

    # útdíj adatok a belföld és PAL-SK
    utdijmappa = config['Utdij-EXCEL']['utdij mappa']
    utdijfile = config['Utdij-EXCEL']['utdij file neve']
    utdij_beresheet = config['Utdij-EXCEL']['utdij bere-munkalap']
    utdij_vacsheet = config['Utdij-EXCEL']['utdij vac-munkalap']
    utdij_sksheet = config['Utdij-EXCEL']['utdij sk-munkalap']
    utdijbalfelso = config['Utdij-EXCEL']['utdij tartomany bal-felso']
    utdijjobbalso = config['Utdij-EXCEL']['utdij tartomany jobb-also']
    kimutatasnev_sheet = config['Utdij-EXCEL']['utdij kimutatasnev-munkalap']
    kimutatasnev_balfelso = config['Utdij-EXCEL']['utdij kimutatasnev tartomany bal-felso']
    kimutatasnev_jobbalso = config['Utdij-EXCEL']['utdij kimutatasnev tartomany jobb-also']
    sapcikk_sheet = config['Utdij-EXCEL']['utdij sapcikkek-munkalap']
    sapcikk_balfelso = config['Utdij-EXCEL']['utdij sapcikkek tartomany bal-felso']
    sapcikk_jobbalso = config['Utdij-EXCEL']['utdij sapcikkek tartomany jobb-also']

    # SAP kódok
    gyvac = config['SAP-Gyarak']['sap vác gyár']
    gybere = config['SAP-Gyarak']['sap beremend gyár']
    gyecser = config['SAP-Gyarak']['sap ecser gyár']

    # útdíj adatok ÖML-SK
    artabla_mappa = config['Fuvardij-Artablak']['artabla mappa']
    skoml_utdijfile = config['Fuvardij-Artablak']['artabla vac-sk-oml file neve']
    skoml_utdijsheet = config['Fuvardij-Artablak']['artabla vac-sk-oml munkalap']
    skoml_utdijbalfelso = config['Fuvardij-Artablak']['artabla vac-sk-oml tartomany bal-felso']
    skoml_utdijjobbalso = config['Fuvardij-Artablak']['artabla vac-sk-oml tartomany jobb-also']
    skoml_utdij2x = config['Fuvardij-Artablak']['artabla vac-sk-oml tartomany utdij2x']
    skoml_utdij1x = config['Fuvardij-Artablak']['artabla vac-sk-oml tartomany utdij1x']

    # útdíj adatok PAL-SK
    skpal_utdijfile = config['Fuvardij-Artablak']['artabla vac-sk-pal file neve']
    skpal_utdijsheet = config['Fuvardij-Artablak']['artabla vac-sk-pal munkalap']
    skpal_utdijbalfelso = config['Fuvardij-Artablak']['artabla vac-sk-pal tartomany bal-felso']
    skpal_utdijjobbalso = config['Fuvardij-Artablak']['artabla vac-sk-pal tartomany jobb-also']
    skpal_utdij2x = config['Fuvardij-Artablak']['artabla vac-sk-pal tartomany utdij2x']
    # ZFT2 1,2x
    skpal_utdij12x = config['Fuvardij-Artablak']['artabla vac-sk-pal tartomany utdij12x']

    saprjohans = config['SAPkodok']['fuvaros-rjohans']
    saphjohans = config['SAPkodok']['fuvaros-hjohans']
    sapkemencepor = config['SAPkodok']['cikk-kemencepor']
    sapspeedline = config['SAPkodok']['fuvaros-speedline']
    sapnordsped = config['SAPkodok']['fuvaros-nordsped']
    sappetranyi = config['SAPkodok']['fuvaros-petranyi']
    # ***** INI fájladatok beolvasása VÉGE

    szlevfajl = str(excelmappa) + str(excelfile)
    # print(szlevfajl)
    logfile.info('Excelfájl (szállítólevelek): %s', str(excelfile))

    helysegfajl = str(helysegmappa) + str(helysegfile)
    # print(helysegfajl)
    logfile.info('Excelfájl (helységnevek): %s', str(helysegfajl))

    utdijfajl = str(utdijmappa) + str(utdijfile)
    # print(utdijfajl)
    logfile.info('Excelfájl (útdíjadatok): %s', str(utdijfajl))

    skoml_artabla = str(artabla_mappa) + str(skoml_utdijfile)
    # print(skoml_artabla)
    logfile.info('Excelfájl (SK-ÖML ártábla): %s', str(skoml_artabla))

    skpal_artabla = str(artabla_mappa) + str(skpal_utdijfile)
    # print(skpal_artabla)
    logfile.info('Excelfájl (SK-PAL ártábla): %s', str(skpal_artabla))

    # Fájl nevének módosítása (szállítólevelek)
    # Fájl végére beillesztésre kerül a napi dátum
    fileneve = excelfile[0:len(excelfile)-5]
    saveasexcelfile = fileneve + "_" + di.mainap("") + ".xlsx"
    pvsaveasexcelfile = 'PV_' + fileneve + "_" + di.mainap("") + ".xlsx"
    logfile.info("Excel fájl (SAVE AS) neve: %s", saveasexcelfile)
    logfile.info("Excel fájl (PIVOT SAVE AS) neve: %s", pvsaveasexcelfile)
    saveas_szlevfajl = str(excelmappa) + str(saveasexcelfile)
    pvsaveas_szlevfajl = str(excelmappa) + str(pvsaveasexcelfile)

    try:
        logfile.info('Excel fájl megnyitása: szállítólevél adatok')
        # adatokat tartalmaézó Excel fájl
        wbook = load_workbook(filename=szlevfajl)
        munkalap = wbook[wbsheetneve]
        # print(munkalap['A1'].value)

        logfile.info('Excel fájl megnyitása: helységnevektábla')
        helyseg_book = load_workbook(filename=helysegfajl, data_only=True)
        helyseg_munkalap = helyseg_book[helysegsheet]
        hrange = helyseg_munkalap[helysegbalfelso: helysegjobbalso]
        # print(hrange)

        logfile.info('Excel fájl megnyitása: útdíjaktábla + kimutatásnevek')
        utdij_book = load_workbook(filename=utdijfajl, data_only=True)
        utberesheet = utdij_book[utdij_beresheet]
        utvacsheet = utdij_book[utdij_vacsheet]
        utsksheet = utdij_book[utdij_sksheet]
        # kimutatásnév munkalap, tartomány
        kimnevsheet = utdij_book[kimutatasnev_sheet]
        kimnevrange = kimnevsheet[kimutatasnev_balfelso: kimutatasnev_jobbalso]
        # SAP cikkek munkalap, tartomány
        sapcikksheet = utdij_book[sapcikk_sheet]
        sapcikkrange = sapcikksheet[sapcikk_balfelso: sapcikk_jobbalso]

        logfile.info('Excel fájl megnyitása: SK-ÖML ártábla')
        skoml_book = load_workbook(filename=skoml_artabla, data_only=True)
        ut_skoml_sheet = skoml_book[skoml_utdijsheet]

        logfile.info('Excel fájl megnyitása: SK-PAL ártábla')
        skpal_book = load_workbook(filename=skpal_artabla, data_only=True)
        ut_skpal_sheet = skpal_book[skpal_utdijsheet]

        maxsor = munkalap.max_row
        logfile.info('Sorok száma (szállítólevelek): %s', str(maxsor))
        maxoszlop = munkalap.max_column
        logfile.info('Oszlopok száma (szállítólevelek): %s', str(maxoszlop))

        #print("sorok: ", str(maxsor))
        #print("oszlopok: ", str(maxoszlop))

        # ***** Excel tábla fejléc

        # Fejléc beállítása az adatbázis mezőnekvekhez igazítva
        fejlecnevek(munkalap)
        logfile.info("Fejlec megnevezés átírása")

        # **** Excel tábla fejléc VÉGE

        # Excel fájl átalakítása

        # Ciklus a soronkénti végig olvasáshoz, íráshoz
        for sor in range(2, maxsor + 1):
            sr = str(sor)

            if type(munkalap['A'+sr].value) == NoneType:
                break

            logfile.info("Szállítólevél sorszáma: %s",
                         str(munkalap['C'+sr].value))

            # Helységnév átalakítása
            arufogado = munkalap['M'+sr].value
            afogado = int(arufogado)
            for j in hrange:
                # vizsgálja, hogy van-e adat a mezőben
                if type(j[0].value) != NoneType:
                    if int(j[0].value) == afogado:
                        munkalap['O'+sr].value = j[1].value
                        logfile.info("Helységnév átalakítva: %s",
                                     str(munkalap['O'+sr].value))
                        break
                else:
                    break
            # **** Helységnév átalakítása VÉGE

            # Incoterms vizsgálata
            # CPT
            if munkalap['E' + sr].value == 'CPT':
                #ÖML - HU - ZF47
                if munkalap['D' + sr].value == '001' and munkalap['P' + sr].value == 'HU' and munkalap['R' + sr].value == 'ZF47':
                    munkalap["AE" + str(
                        sr)].value = munkalap["AH" + sr].value - munkalap["AG" + sr].value
                    logfile.info("ÖML - HU - ZF47 fuvardíj: %s",
                                 str(munkalap['AE'+sr].value))

                #PAL - HU - ZF47
                if munkalap['D' + sr].value == '006' and munkalap['P' + sr].value == 'HU' and munkalap['R' + sr].value == 'ZF47':
                    munkalap["AE" + str(
                        sr)].value = munkalap["AH" + sr].value - munkalap["AG" + sr].value
                    logfile.info("PAL - HU - ZF47 fuvardíj: %s",
                                 str(munkalap['AE'+sr].value))

                #ÖML - HU - ZF49
                if munkalap['D' + sr].value == '001' and munkalap['P' + sr].value == 'HU' and munkalap['R' + sr].value == 'ZF49':
                    # JOHANS rövidkód
                    fuvarozokod = munkalap['I'+sr].value
                    if fuvarozokod == saprjohans:
                        if munkalap['B'+sr].value == gyvac:
                            utrange = utvacsheet[utdijbalfelso: utdijjobbalso]
                        if munkalap['B'+sr].value == gybere:
                            utrange = utberesheet[utdijbalfelso: utdijjobbalso]

                        for i in utrange:
                            if type(i[0].value) != NoneType:
                                szlevhelyseg = str(munkalap['O'+sr].value)
                                uthelyseg = str(i[1].value)
                                if szlevhelyseg.upper() == uthelyseg.upper():
                                    munkalap['AG'+sr].value = i[9].value
                                    logfile.info(
                                        "ÖML - HU - ZF47 JOHANS (rövid) útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                            else:
                                break

                        munkalap["AE" + str(
                            sr)].value = munkalap["AH" + sr].value - munkalap["AG" + sr].value
                    # **** JOHANS rövidkód VÉGE

                    # JOHANS hosszúkód
                    fuvarozokod = munkalap['I'+sr].value
                    if fuvarozokod == saphjohans:
                        if munkalap['B'+sr].value == gyvac:
                            utrange = utvacsheet[utdijbalfelso: utdijjobbalso]
                        if munkalap['B'+sr].value == gybere:
                            utrange = utberesheet[utdijbalfelso: utdijjobbalso]

                        for i in utrange:
                            if type(i[0].value) != NoneType:
                                szlevhelyseg = str(munkalap['O'+sr].value)
                                uthelyseg = str(i[1].value)
                                if szlevhelyseg.upper() == uthelyseg.upper():
                                    munkalap['AG'+sr].value = i[9].value
                                    logfile.info(
                                        "ÖML - HU - ZF47 JOHANS (rövid) útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                            else:
                                break

                        munkalap["AE" + str(
                            sr)].value = munkalap["AH" + sr].value - munkalap["AG" + sr].value
                    # **** JOHANS hosszúkód VÉGE

                    # Kemencepor
                    termek = munkalap['G'+sr].value
                    if termek == sapkemencepor:
                        if munkalap['B'+sr].value == gyvac:
                            utrange = utvacsheet[utdijbalfelso: utdijjobbalso]
                        if munkalap['B'+sr].value == gybere:
                            utrange = utberesheet[utdijbalfelso: utdijjobbalso]

                        for i in utrange:
                            if type(i[0].value) != NoneType:
                                szlevhelyseg = str(munkalap['O'+sr].value)
                                uthelyseg = str(i[1].value)
                                if szlevhelyseg.upper() == uthelyseg.upper():
                                    munkalap['AG'+sr].value = i[9].value * 2
                                    logfile.info(
                                        "ÖML - HU - ZF47 KEMENCEPOR útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                            else:
                                break

                        munkalap["AE" + str(
                            sr)].value = munkalap["AH" + sr].value - munkalap["AG" + sr].value
                    # **** Kemencepor VÉGE
                # **** ÖML - HU - ZF49 VÉGE

                # ÖML - SK - ZF49
                if munkalap['D' + sr].value == '001' and munkalap['P' + sr].value == 'SK' and munkalap['R' + sr].value == 'ZF49':
                    utrange = ut_skoml_sheet[skoml_utdijbalfelso: skoml_utdijjobbalso]

                    fuvarozokod = munkalap['I'+sr].value
                    arufogado = munkalap['M'+sr].value
                    afogado = int(arufogado)
                    for j in utrange:
                        # vizsgálja, hogy van-e adat a mezőben
                        if type(j[0].value) != NoneType:
                            if int(j[0].value) == afogado:
                                if fuvarozokod[0:4] == '1832':
                                    # oszlopok száma nullával kezdődik, így a táblában lévő
                                    # sorszámból egyet le kell vonni
                                    #munkalap['AG'+sr].value = j[67].value
                                    munkalap['AG' +
                                             sr].value = j[int(skoml_utdij1x)].value
                                    logfile.info(
                                        "ÖML - HU - ZF49 1x útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                                else:
                                    #munkalap['AG'+sr].value = j[66].value
                                    munkalap['AG' +
                                             sr].value = j[int(skoml_utdij2x)].value
                                    logfile.info(
                                        "ÖML - HU - ZF49 2x útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                        else:
                            break

                    munkalap["AE" + str(
                        sr)].value = munkalap["AH" + sr].value - munkalap["AG" + sr].value
                # **** ÖML - SK - ZF49 - hosszúkód VÉGE

                # PAL - SK - ZF49
                if munkalap['D' + sr].value == '006' and munkalap['P' + sr].value == 'SK' and munkalap['R' + sr].value == 'ZF49':
                    utrange = ut_skpal_sheet[skpal_utdijbalfelso: skpal_utdijjobbalso]

                    fuvarozokod = munkalap['I'+sr].value
                    arufogado = munkalap['M'+sr].value
                    afogado = int(arufogado)
                    for j in utrange:
                        # vizsgálja, hogy van-e adat a mezőben
                        if type(j[0].value) != NoneType:
                            if int(j[0].value) == afogado:
                                if fuvarozokod == sapspeedline:
                                    munkalap['AG'+sr].value = 0
                                    logfile.info(
                                        "PAL - HU - ZF49 SpeedLine útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                                if fuvarozokod == sapnordsped:
                                    munkalap['AG'+sr].value = 0
                                    logfile.info(
                                        "PAL - HU - ZF49 Nordsped útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                                if fuvarozokod == sappetranyi:
                                    # oszlopok száma nullával kezdődik, így a táblában lévő
                                    # sorszámból egyet le kell vonni
                                    #munkalap['AG' + sr].value = j[22].value
                                    munkalap['AG' +
                                             sr].value = j[int(skpal_utdij2x)].value
                                    logfile.info(
                                        "PAL - HU - ZF49 Petrányi 2x útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                                else:
                                    #munkalap['AG' + sr].value = j[21].value
                                    munkalap['AG' +
                                             sr].value = j[int(skpal_utdij12x)].value
                                    logfile.info(
                                        "PAL - HU - ZF49 1,2x útdíj: %s", str(munkalap['AG'+sr].value))
                                    break
                        else:
                            break

                    munkalap["AE" + str(
                        sr)].value = munkalap["AH" + sr].value - munkalap["AG" + sr].value
                # **** PAL - SK - ZF49 - hosszúkód VÉGE
            # **** CPT VÉGE

            # EXW
            elif munkalap['E' + sr].value == 'EXW':
                munkalap["AE" + sr].value = 0  # nettó fuvardíj
                munkalap["AG" + sr].value = 0  # útdíj
                munkalap["AH" + sr].value = 0  # fuvar- + útdíj
            # FCA incoterms esetén átírás EXW-re
            elif munkalap['E' + sr].value == 'FCA':
                munkalap['E' + sr].value = 'EXW'
                munkalap["AE" + sr].value = 0  # nettó fuvardíj
                munkalap["AG" + sr].value = 0  # útdíj
                munkalap["AH" + sr].value = 0  # fuvar- + útdíj
            # Incoterms a fentiek közül egyiksem
            else:
                munkalap["AE" + sr].value = 0
                munkalap["AG" + sr].value = 0
                munkalap["AH" + sr].value = 0

            # **** Incoterms vizsgálata VÉGE

            # EUR árfolyam konvertálása számra

            # áttárolás/értékesítés/visszavét beállítása
            megrendelokod = munkalap['K'+sr].value
            if str(megrendelokod) == '18160032':
                munkalap['AM'+sr].value = 'Áttárolás'
                logfile.info("Áttárolás beállítva: %s",
                             str(munkalap['AM'+sr].value))
            else:
                munkalap['AM'+sr].value = 'Értékesítés'
                logfile.info("Értékesítés beállítva: %s",
                             str(munkalap['AM'+sr].value))
            # **** áttárolás/értékesítés/visszavét beállítása

            # kimutatás nevek beállítása, amelyik fuvarozónak van SAP kódja
            vanfuvarozokod = False
            fuvarozokod = munkalap['I'+sr].value
            for i in kimnevrange:
                if type(i[0].value) != NoneType:
                    if str(fuvarozokod) == str(i[0].value):
                        munkalap['AL' + sr].value = i[2].value
                        logfile.info("Kimutatásnév beállítva: %s", str(
                            munkalap['AL'+sr].value))
                        vanfuvarozokod = True
                        break
                else:
                    break
            if vanfuvarozokod == False:
                munkalap['AL' + sr].value = 'ism. fuvarozó'
                logfile.info("Ismeretlen fuvarozó beállítva: %s",
                             str(munkalap['AL'+sr].value))
            # **** kimutatás nevek beállítása VÉGE

            # termékfajta (cement, kó, kemencepor, stb.) beállítása
            sapcikk_kod = munkalap['G'+sr].value
            for i in sapcikkrange:
                if type(i[0].value) != NoneType:
                    if str(sapcikk_kod) == str(i[0].value):
                        munkalap['AN'+sr].value = i[5].value
                        logfile.info("Termékcsoport beállítva: %s", str(
                            munkalap['AN'+sr].value))
                        break
                else:
                    break
            # **** termékfajta (cement, kó, kemencepor, stb.) beállítása

        # **** Ciklus a soronkénti végig olvasáshoz, íráshoz

        # Excel fájlok bezárása
        wbook.save(saveas_szlevfajl)
        wbook.close()
        logfile.info('Excel fájl elmentve, bezárva (szállítólevél adatok).')

        helyseg_book.close()
        logfile.info('Excel fájl elmentve, bezárva (helységnevek).')

        utdij_book.close()
        logfile.info('Excel fájl elmentve, bezárva (útdijak).')

        skoml_book.close()
        logfile.info('Excel fájl elmentve, bezárva (SK-ÖML ártábla).')

        skpal_book.close()
        logfile.info('Excel fájl elmentve, bezárva (SK-PAL ártábla).')

        # **** Excel fájl átalakítása VÉGE

        # Pivot Table-k összeállítása

        # mpt.pivottabla(prgneve, saveas_szlevfajl, pvsaveas_szlevfajl, logfile)

        # **** Pivot Table-k összeállítása VÉGE

    # **** TRY VÉGE

    except FileNotFoundError:
        logfile.error("Az XLSX fájl nem található: %s", str(szlevfajl))
        logfile.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Az XLSX Fájl nem található!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)
    except Exception as merror:
        logfile.error(
            'Ismeretlen hiba típusa, leírás: %s: %s', str(type(merror)), str(merror))
        logfile.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Ismeretlen hiba!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)

    logfile.info('Exel forrásfájl feldolgozása befejeződött')

    # átalakított adatokat tartalmazó Excel fájl neve (elérési úttal együtt)
    return saveas_szlevfajl


def fejlecnevek(rs):
    """
    Fejlécnevek átírása, az adatbázis mezőkhöz igazítva
    Készült: 2022.06.16

    Paraméter:
        rs - Excelmunkalap változónév

    Utolsó módosítás dátuma: 2022.07.20
    Verzió: 01
    """
    rs["A1"].value = "ErtSzerv"
    rs["A1"].fill = PatternFill(bgColor='ffffff')
    rs["B1"].value = "Gyar"
    rs["B1"].fill = PatternFill(bgColor='ffffff')
    rs["C1"].value = "SzlevSzama"
    rs["C1"].fill = PatternFill(bgColor='ffffff')
    rs["D1"].value = "Csomagolas"
    rs["D1"].fill = PatternFill(bgColor='ffffff')
    rs["E1"].value = "Incoterms"
    rs["E1"].fill = PatternFill(bgColor='ffffff')
    rs["F1"].value = "Tetel"
    rs["F1"].fill = PatternFill(bgColor='ffffff')
    rs["G1"].value = "AnyagKod"
    rs["G1"].fill = PatternFill(bgColor='ffffff')
    rs["H1"].value = "Rendszam"
    rs["H1"].fill = PatternFill(bgColor='ffffff')
    rs["I1"].value = "FuvarozoKod"
    rs["I1"].fill = PatternFill(bgColor='ffffff')
    rs["J1"].value = "FuvarozoNeve"
    rs["J1"].fill = PatternFill(bgColor='ffffff')
    rs["K1"].value = "MegrendeloKod"
    rs["K1"].fill = PatternFill(bgColor='ffffff')
    rs["L1"].value = "MegrendeloNeve"
    rs["L1"].fill = PatternFill(bgColor='ffffff')
    rs["M1"].value = "ArufogadoKod"
    rs["M1"].fill = PatternFill(bgColor='ffffff')
    rs["N1"].value = "ArufogadoNeve"
    rs["N1"].fill = PatternFill(bgColor='ffffff')
    rs["O1"].value = "Helyseg"
    rs["O1"].fill = PatternFill(bgColor='ffffff')
    rs["P1"].value = "Orszag"
    rs["P1"].fill = PatternFill(bgColor='ffffff')
    rs["Q1"].value = "VevoKorzet"
    rs["Q1"].fill = PatternFill(bgColor='ffffff')
    rs["R1"].value = "KondicioFajta"
    rs["R1"].fill = PatternFill(bgColor='ffffff')
    rs["S1"].value = "RendelesSzama"
    rs["S1"].fill = PatternFill(bgColor='ffffff')
    rs["T1"].value = "SzamlaSzama"
    rs["T1"].fill = PatternFill(bgColor='ffffff')
    rs["U1"].value = "SzlevDatum"
    rs["U1"].fill = PatternFill(bgColor='ffffff')
    rs["V1"].value = "Tonna"
    rs["V1"].fill = PatternFill(bgColor='ffffff')
    rs["W1"].value = "TonnaMertEgys"
    rs["W1"].fill = PatternFill(bgColor='ffffff')
    rs["X1"].value = "Tavolsag"
    rs["X1"].fill = PatternFill(bgColor='ffffff')
    rs["Y1"].value = "TavolsagMertEgys"
    rs["Y1"].fill = PatternFill(bgColor='ffffff')
    rs["Z1"].value = "SzlaNettoErtek"
    rs["Z1"].fill = PatternFill(bgColor='ffffff')
    rs["AA1"].value = "SzlaPenznem"
    rs["AA1"].fill = PatternFill(bgColor='ffffff')
    rs["AB1"].value = "ATKm"
    rs["AB1"].fill = PatternFill(bgColor='ffffff')
    rs["AC1"].value = "FuvarEgysAr"
    rs["AC1"].fill = PatternFill(bgColor='ffffff')
    rs["AD1"].value = "FuvarPenznem"
    rs["AD1"].fill = PatternFill(bgColor='ffffff')
    rs["AE1"].value = "FuvardijNetto"
    rs["AE1"].fill = PatternFill(bgColor='ffffff')
    rs["AF1"].value = "UtdijKondicio"
    rs["AF1"].fill = PatternFill(bgColor='ffffff')
    rs["AG1"].value = "Utdij"
    rs["AG1"].fill = PatternFill(bgColor='ffffff')
    rs["AH1"].value = "FuvarUtdijBrutto"
    rs["AH1"].fill = PatternFill(bgColor='ffffff')
    rs["AI1"].value = "EURArfolyam"
    rs["AI1"].fill = PatternFill(bgColor='ffffff')
    rs["AJ1"].value = "RendelesTipus"
    rs["AJ1"].fill = PatternFill(bgColor='ffffff')
    rs["AK1"].value = "NettoFuvardij"
    rs["AK1"].fill = PatternFill(bgColor='ffffff')
    rs["AL1"].value = "Kimutatasnev"
    rs["AL1"].fill = PatternFill(bgColor='ffffff')
    rs["AM1"].value = "Attarolas"
    rs["AM1"].fill = PatternFill(bgColor='ffffff')
    rs["AN1"].value = "TermekCsoport"
    rs["AN1"].fill = PatternFill(bgColor='ffffff')
