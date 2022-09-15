"""
SAP szállítólevelet (havi zárás) tartalmazó, Excel táblázat adatainak 
felvezetése az SQL adatbázisba.

Készült: 2022.08.05

Utolsó módosítás dátuma: 2022.09.15
verzió: 02

"""

import sys
import ctypes
from types import NoneType
from openpyxl import load_workbook
from datetime import datetime

import read_config as readcfg
import datumido as di
import adatbconnect as dbconnect


def urescella(cella):
    # Excel munkalap cellájának vizsgálata, hogy üres-e
    # ha igen, üres karakterlácot ad vissza, egyébként a celle értékét

    if type(cella) == NoneType:
        megjegyzes = ''
    else:
        megjegyzes = cella

    return megjegyzes


def SAPSzlev_feltoltese(excelfile, inifajl, logfile):
    '''
    Excel tábla tartalmának betöltése az SQL adatbázisba
    ODBC kapcsolattal.

    Paraméterek:
    excelfile = ez a fájl tartalmazza a betöltendő adatokat
    inifajl = INI fájlban lévő adatok ebből a fájlból olvashatók ki
    logile = LOG fájl írása

    Utolső módosítás dátuma: 2022.09.15
    '''

    # INI fájl megnyitása
    config = readcfg.read_config(inifajl)

    logfile.info('Adatbázis műveletek: SAPSzallitolevelek')
    prgneve = config['EXE-File']['exe-file neve']

    sapszlev_sheet = config['SAP-Excel']['sapexcel munkalap']

    odbc_driver = config['SQL-Server']['odbc driver']
    logfile.info('ODBC driver: %s', str(odbc_driver))
    sqlszerver = config['SQL-Server']['sql server name']
    logfile.info('SQL szerver: %s', str(sqlszerver))
    sqladatb = config['SQL-Server']['sql database name']
    logfile.info('Adatbázis: %s', str(sqladatb))
    sqlSAPSzlev = config['SQL-Server']['sql szallitolevel adattábla']
    logfile.info('Adatbázis-tábla: %s', str(sqlSAPSzlev))

    logfile.info('Excelfájl (SAP szállítólevelek): %s', str(excelfile))

    FelvezetesDatuma = di.mainap("kotojel")

    try:

        logfile.info('Excel fájl megnyitása a SAP szállítólevelekhez')
        szlev_book = load_workbook(filename=excelfile, data_only=True)
        sapszlevsheet = szlev_book[sapszlev_sheet]

        logfile.info('ODBC kapcsolat megnyitása')
        kapcs = dbconnect.ODBC_Kapcsolat(odbc_driver, sqlszerver, sqladatb)
        kapcscursor = kapcs.cursor()

        maxsor = sapszlevsheet.max_row
        logfile.info('Sorok száma (szállítólevelek): %s', str(maxsor))
        maxoszlop = sapszlevsheet.max_column
        logfile.info('Oszlopok száma (szállítólevelek): %s', str(maxoszlop))

        # adatok törlése
        '''
        delete_sql = "DELETE FROM [dbo].[Szallitolevelek]"
        kapcscursor.execute(delete_sql)
        kapcs.commit()
        '''

        for sor in range(2, maxsor+1):
            aktsor = str(sor)

            if type(sapszlevsheet['A'+aktsor].value) == NoneType:
                break

            logfile.info("Szállítólevélszáma: %s", str(
                sapszlevsheet['C'+aktsor].value))

            mrendszam = urescella(sapszlevsheet['H' + aktsor].value)
            logfile.info('Rendszám: %s', str(mrendszam))
            mfuvarozokod = urescella(sapszlevsheet['I' + aktsor].value)
            logfile.info('Fuvarozó kód: %s', str(mfuvarozokod))
            mfuvarozoneve = urescella(sapszlevsheet['J' + aktsor].value)
            logfile.info('Fuvarozó neve: %s', str(mfuvarozoneve))
            mkondiciofajta = urescella(sapszlevsheet['R' + aktsor].value)
            logfile.info('Kondíciófajta: %s', str(mkondiciofajta))
            mszamlaszama = urescella(sapszlevsheet['T' + aktsor].value)
            logfile.info('Számlaszáma: %s', str(mszamlaszama))

            mnettofuvardij = sapszlevsheet['AH' +
                                           aktsor].value - sapszlevsheet['AG' + aktsor].value
            logfile.info('Nettó fuvardíj: %s', str(mnettofuvardij))

            #insert_sql=("INSERT INTO Szallitolevelek ([FuvardijNetto]) VALUES(?)")
            #adat=(sapszlevsheet['AE' + aktsor].value)
            #print(sapszlevsheet['AE' + aktsor].value)
            #print(type(sapszlevsheet['AE' + aktsor].value))

            insert_sql = ("INSERT INTO Szallitolevelek ([ErtSzerv], [Gyar], [SzlevSzama], [Csomagolas], [Incoterms], [Tetel], [AnyagKod], [Rendszam], [FuvarozoKod]," +
                          " [FuvarozoNeve], [MegrendeloKod], [MegrendeloNeve], [ArufogadoKod], [ArufogadoNeve], [Helyseg], [Orszag], [VevoKorzet], [KondicioFajta]," +
                          " [RendelesSzama], [SzamlaSzama], [SzlevDatum], [Tonna], [TonnaMertEgys], [Tavolsag], [TavolsagMertEgys], [SzlaNettoErtek], [SzlaPenznem]," +
                          " [ATKm], [FuvarEgysAr], [FuvarPenznem], [FuvardijNetto], [UtdijKondicio], [Utdij], [FuvarUtdijBrutto], [EURArfolyam], [RendelesTipus]," +
                          " [NettoFuvardij], [Kimutatasnev], [Attarolas], [TermekCsoport], [RogzitesDatuma])" +
                          " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)")

            # ErtSzerv
            adat = (sapszlevsheet['A' + aktsor].value,
                    # Gyar
                    sapszlevsheet['B' + aktsor].value,
                    # SzlevSzama
                    sapszlevsheet['C' + aktsor].value,
                    # Csomagolas
                    sapszlevsheet['D' + aktsor].value,
                    # Incoterms
                    sapszlevsheet['E' + aktsor].value,
                    # Tetel
                    sapszlevsheet['F' + aktsor].value,
                    # AnyagKod
                    sapszlevsheet['G' + aktsor].value,
                    # Rendszam
                    mrendszam,
                    # FuvarozoKod
                    mfuvarozokod,
                    # FuvarozoNeve
                    mfuvarozoneve,
                    # MerendeloKod
                    sapszlevsheet['K' + aktsor].value,
                    # MegrendeloNeve
                    sapszlevsheet['L' + aktsor].value,
                    # ArufogadoKod
                    sapszlevsheet['M' + aktsor].value,
                    # ArufogadoNeve
                    sapszlevsheet['N' + aktsor].value,
                    # Helyseg
                    sapszlevsheet['O' + aktsor].value,
                    # Orszag
                    sapszlevsheet['P' + aktsor].value,
                    # VevoKorzet
                    sapszlevsheet['Q' + aktsor].value,
                    # KondicioFajta
                    mkondiciofajta,
                    # RendelesSzama
                    sapszlevsheet['S' + aktsor].value,
                    # SzamlaSzama
                    mszamlaszama,
                    # SzlevDatum
                    sapszlevsheet['U' + aktsor].value,
                    # Tonna
                    sapszlevsheet['V' + aktsor].value,
                    # TonnaMertEgys
                    sapszlevsheet['W' + aktsor].value,
                    # Tavolsag
                    sapszlevsheet['X' + aktsor].value,
                    # TavolsagMertEgys
                    sapszlevsheet['Y' + aktsor].value,
                    # SzlaNettoErtek
                    sapszlevsheet['Z' + aktsor].value,
                    # SzlaPenznem
                    sapszlevsheet['AA' + aktsor].value,
                    # ATKm
                    sapszlevsheet['AB' + aktsor].value,
                    # FuvarEgyyAr
                    sapszlevsheet['AC' + aktsor].value,
                    # FuvarPenznem
                    sapszlevsheet['AD' + aktsor].value,
                    # FuvardijNetto
                    sapszlevsheet['AO' + aktsor].value,
                    # UtdijKondicio
                    sapszlevsheet['AF' + aktsor].value,
                    # Utdij
                    sapszlevsheet['AG' + aktsor].value,
                    # FuvarUtdijBrutto
                    sapszlevsheet['AH' + aktsor].value,
                    # EURArfolyam
                    sapszlevsheet['AI' + aktsor].value,
                    # RendelesTipus
                    sapszlevsheet['AJ' + aktsor].value,
                    # NettoFuvardíj
                    sapszlevsheet['AK' + aktsor].value,
                    # Kimutatasnev
                    sapszlevsheet['AL' + aktsor].value,
                    # Attarolas
                    sapszlevsheet['AM' + aktsor].value,
                    # TermekCsoport
                    sapszlevsheet['AN' + aktsor].value,
                    # RogzitesDatuma
                    FelvezetesDatuma
                    )
            kapcscursor.execute(insert_sql, adat)
            kapcs.commit()

        # ***** for sor in range(2, maxsor+1): VÉGE

    except Exception as merror:

        # Adatbázis műveletek visszavonása, a hiba miatt
        kapcs.rollback()

        # Logfájlba írás
        logfile.error(
            'Ismeretlen hiba típusa, leírás: %s: %s', str(type(merror)), str(merror))
        logfile.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Ismeretlen hiba!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)

    # ODBC kapcsolat bezárása
    kapcs.close()
    logfile.info(
        'SAPSzallitolevelek feltöltöltve, adatbáziskapcsolat bezárása.')
    # ***** SAPCikkek_feltoltese VÉGE
