"""
SAP cikkeket tartalmazó, Excel táblázat adatainak 
felvezetése az SQL adatbázisba.

Készült: 2022.08.04

Utolsó módosítás dátuma: 2022.09.15
verzió: 02

"""

import sys
import ctypes
from types import NoneType
from openpyxl import load_workbook

import read_config as readcfg
import datumido as di
import adatbconnect as dbconnect


def SAPCikkek_feltoltese(inifajl, logfile):

    # INI fájl megnyitása
    config = readcfg.read_config(inifajl)

    logfile.info('Adatbázis műveletek: SAPCikkek')
    prgneve = config['EXE-File']['exe-file neve']

    utdijmappa = config['Utdij-EXCEL']['utdij mappa']
    utdijfile = config['Utdij-EXCEL']['utdij file neve']
    sapcikk_sheet = config['Utdij-EXCEL']['utdij sapcikkek-munkalap']

    odbc_driver = config['SQL-Server']['odbc driver']
    logfile.info('ODBC driver: %s', str(odbc_driver))
    sqlszerver = config['SQL-Server']['sql server name']
    logfile.info('SQL szerver: %s', str(sqlszerver))
    sqladatb = config['SQL-Server']['sql database name']
    logfile.info('Adatbázis: %s', str(sqladatb))
    sqlSAPCikkek = config['SQL-Server']['sql sap cikkek adattabla']
    logfile.info('Adatbázis-tábla: %s', str(sqlSAPCikkek))

    utdijfajl = str(utdijmappa) + str(utdijfile)
    logfile.info('Excelfájl (SAP cikkek): %s', str(utdijfajl))

    FelvezetesDatuma = di.mainap("kotojel")

    try:

        logfile.info('Excel fájl megnyitása a SAP cikkekhez')
        utdij_book = load_workbook(filename=utdijfajl, data_only=True)
        sapcikksheet = utdij_book[sapcikk_sheet]

        logfile.info('ODBC kapcsolat megnyitása')
        kapcs = dbconnect.ODBC_Kapcsolat(odbc_driver, sqlszerver, sqladatb)
        kapcscursor = kapcs.cursor()

        # Táblában lévő adatok törlése (DELETE ALL)
        logfile.info('SAPCikkek tábla adatainak törlése')
        kapcscursor.execute("DELETE FROM " + str(sqlSAPCikkek))
        kapcscursor.commit()

        maxsor = sapcikksheet.max_row
        logfile.info('Sorok száma (szállítólevelek): %s', str(maxsor))
        maxoszlop = sapcikksheet.max_column
        logfile.info('Oszlopok száma (szállítólevelek): %s', str(maxoszlop))

        for sor in range(2, maxsor+1):
            aktsor = str(sor)

            if type(sapcikksheet['A'+aktsor].value) == NoneType:
                break

            logfile.info("Cikkszám: %s", str(sapcikksheet['A'+aktsor].value))

            if type(sapcikksheet['I'+aktsor].value) == NoneType:
                megjegyzes = ''
            else:
                megjegyzes = sapcikksheet['I'+aktsor].value

            sql = (
                "INSERT INTO SAPCikkek VALUES( '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')"
                % (
                    sapcikksheet['A'+aktsor].value,
                    sapcikksheet['B'+aktsor].value,
                    sapcikksheet['C'+aktsor].value,
                    sapcikksheet['D'+aktsor].value,
                    sapcikksheet['E'+aktsor].value,
                    sapcikksheet['F'+aktsor].value,
                    sapcikksheet['G'+aktsor].value,
                    sapcikksheet['H'+aktsor].value,
                    megjegyzes,
                    FelvezetesDatuma,
                )
            )

            kapcscursor.execute(sql)
            kapcs.commit()

        # ***** for sor in range(2, maxsor+1): VÉGE

    except Exception as merror:
        logfile.error(
            'Ismeretlen hiba típusa, leírás: %s: %s', str(type(merror)), str(merror))
        logfile.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Ismeretlen hiba!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)

    # ODBC kapcsolat bezárása
    kapcs.close()
    logfile.info('SAPCikkek feltöltöltve, adatbáziskapcsolat bezárása.')
    # ***** SAPCikkek_feltoltese VÉGE
