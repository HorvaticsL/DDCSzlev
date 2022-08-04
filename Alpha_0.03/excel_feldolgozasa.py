"""
Excel fájl tartalmának feldolgozása
A fájlban lévő adatokat úgy kell átalakítani, hogy
az adatbázisba lementhetők legyenek - helységnevek, fuvar-, útdíjak, stb.
Készült: 2022.06.13

Utolsó módosítás dátuma: 2022.07.12
verzió: 02

"""
#import openpyxl
from encodings.utf_8 import encode
from imp import load_dynamic
import sys
import ctypes
from types import NoneType
from openpyxl import Workbook, load_workbook
#from openpyxl.utils import get_column_letter

import ini_fajl as inif
import naplozas
import datumido as di


def excelfajl_modositas(initomb):
    # Utolsó módosítás dátuma: 2022.07.12
    # Naplózás beállítása
    excelnaplo = naplozas.naplolog()
    excelnaplo.info('Exel forrásfájl feldolgozása elindult')

    # INI fájladatok beolvasása
    prgneve = inif.initomb_eleme(initomb, 0)
    # szállítólevelek forrás fájl
    excelfile = inif.initomb_eleme(initomb, 3)
    excelmappa = inif.initomb_eleme(initomb, 4)
    wbsheetneve = inif.initomb_eleme(initomb, 5)

    # helységnevek
    helysegmappa = inif.initomb_eleme(initomb, 6)
    helysegfile = inif.initomb_eleme(initomb, 7)
    helysegsheet = inif.initomb_eleme(initomb, 8)
    helysegbalfelso = inif.initomb_eleme(initomb, 9)
    helysegjobbalso = inif.initomb_eleme(initomb, 10)

    # útdíj adatok a belföld és PAL-SK
    utdijmappa = inif.initomb_eleme(initomb, 11)
    utdijfile = inif.initomb_eleme(initomb, 12)
    utdij_beresheet = inif.initomb_eleme(initomb, 13)
    utdij_vacsheet = inif.initomb_eleme(initomb, 14)
    utdij_sksheet = inif.initomb_eleme(initomb, 15)
    utdijbalfelso = inif.initomb_eleme(initomb, 16)
    utdijjobbalso = inif.initomb_eleme(initomb, 17)

    # SAP kódok
    gyvac = inif.initomb_eleme(initomb, 18)
    gybere = inif.initomb_eleme(initomb, 19)
    gyecser = inif.initomb_eleme(initomb, 20)

    # útdíj adatok ÖML-SK
    artabla_mappa = inif.initomb_eleme(initomb, 21)
    skoml_utdijfile = inif.initomb_eleme(initomb, 24)
    skoml_utdijsheet = inif.initomb_eleme(initomb, 43)
    skoml_utdijbalfelso = inif.initomb_eleme(initomb, 25)
    skoml_utdijjobbalso = inif.initomb_eleme(initomb, 26)

    # útdíj adatok PAL-SK
    skpal_utdijfile = inif.initomb_eleme(initomb, 27)
    skpal_utdijsheet = inif.initomb_eleme(initomb, 44)
    skpal_utdijbalfelso = inif.initomb_eleme(initomb, 28)
    skpal_utdijjobbalso = inif.initomb_eleme(initomb, 29)

    sapjohans = inif.initomb_eleme(initomb, 30)
    sapkemencepor = inif.initomb_eleme(initomb, 31)
    sapspeedline = inif.initomb_eleme(initomb, 32)
    sapnordsped = inif.initomb_eleme(initomb, 33)
    sappetranyi = inif.initomb_eleme(initomb, 34)
    # INI fájladatok beolvasása VÉGE

    szlevfajl = str(excelmappa) + str(excelfile)
    # print(szlevfajl)
    excelnaplo.info('Excelfájl (szállítólevelek): %s', str(excelfile))

    helysegfajl = str(helysegmappa) + str(helysegfile)
    # print(helysegfajl)
    excelnaplo.info('Excelfájl (helységnevek): %s', str(helysegfajl))

    utdijfajl = str(utdijmappa) + str(utdijfile)
    # print(utdijfajl)
    excelnaplo.info('Excelfájl (útdíjadatok): %s', str(utdijfajl))

    skoml_artabla = str(artabla_mappa) + str(skoml_utdijfile)
    # print(skoml_artabla)
    excelnaplo.info('Excelfájl (SK-ÖML ártábla): %s', str(skoml_artabla))

    skpal_artabla = str(artabla_mappa) + str(skpal_utdijfile)
    # print(skpal_artabla)
    excelnaplo.info('Excelfájl (SK-PAL ártábla): %s', str(skpal_artabla))

    # Fájl nevének módosítása (szállítólevelek)
    # Fájl végére beillesztésre kerül a napi dátum
    fileneve = excelfile[0:len(excelfile)-5]
    saveasexcelfile = fileneve + "_" + di.mainap("") + ".xlsx"
    excelnaplo.info("Excel fájl (SAVE AS) neve: %s", saveasexcelfile)
    saveas_szlevfajl = str(excelmappa) + str(saveasexcelfile)

    try:
        excelnaplo.info('Excel fájl megnyitása: szállítólevél adatok')
        # adatokat tartalmaézó Excel fájl
        wbook = load_workbook(filename=szlevfajl)
        munkalap = wbook[wbsheetneve]
        # print(munkalap['A1'].value)

        excelnaplo.info('Excel fájl megnyitása: helységnevektábla')
        helyseg_book = load_workbook(filename=helysegfajl, data_only=True)
        helyseg_munkalap = helyseg_book[helysegsheet]
        hrange = helyseg_munkalap[helysegbalfelso: helysegjobbalso]
        # print(hrange)

        excelnaplo.info('Excel fájl megnyitása: útdíjaktábla')
        utdij_book = load_workbook(filename=utdijfajl, data_only=True)
        utberesheet = utdij_book[utdij_beresheet]
        utvacsheet = utdij_book[utdij_vacsheet]
        utsksheet = utdij_book[utdij_sksheet]

        excelnaplo.info('Excel fájl megnyitása: SK-ÖML ártábla')
        skoml_book = load_workbook(filename=skoml_artabla, data_only=True)
        ut_skoml_sheet = skoml_book[skoml_utdijsheet]

        excelnaplo.info('Excel fájl megnyitása: SK-PAL ártábla')
        skpal_book = load_workbook(filename=skpal_artabla, data_only=True)
        ut_skpal_sheet = skpal_book[skpal_utdijsheet]

        maxsor = munkalap.max_row
        excelnaplo.info('Sorok száma (szállítólevelek): %s', str(maxsor))
        maxoszlop = munkalap.max_column
        excelnaplo.info('Oszlopok száma (szállítólevelek): %s', str(maxoszlop))

        #print("sorok: ", str(maxsor))
        #print("oszlopok: ", str(maxoszlop))

        # ***** Excel tábla fejléc

        # Fejléc beállítása az adatbázis mezőnekvekhez igazítva
        fejlecnevek(munkalap)
        excelnaplo.info("Fejlec megnevezés átírása")

        # **** Excel tábla fejléc VÉGE

        # **** Excel fájl átalakítása

        # Ciklus a soronkénti végig olvasáshoz, íráshoz
        for sor in range(2, maxsor):
            sr = str(sor)

            # Helységnév átalakítása
            arufogado = munkalap['M'+sr].value
            afogado = int(arufogado)
            for j in hrange:
                # vizsgálja, hogy van-e adat a mezőben
                if type(j[0].value) != NoneType:
                    if int(j[0].value) == afogado:
                        munkalap['O'+sr].value = j[1].value
                        break
                else:
                    break
            # Helységnév átalakítása VÉGE

            # Incoterms vizsgálata
            # CPT
            if munkalap['E' + sr].value == 'CPT':
                #ÖML - HU - ZF47
                if munkalap['D' + sr].value == '001' and munkalap['P' + sr].value == 'HU' and munkalap['R' + sr].value == 'ZF47':
                    munkalap["AE" + str(
                        sr)].value = f'={munkalap["AH" + sr].value - munkalap["AG" + sr].value}'

                #PAL - HU - ZF47
                if munkalap['D' + sr].value == '006' and munkalap['P' + sr].value == 'HU' and munkalap['R' + sr].value == 'ZF47':
                    munkalap["AE" + str(
                        sr)].value = f'={munkalap["AH" + sr].value - munkalap["AG" + sr].value}'

                #ÖML - HU - ZF49
                if munkalap['D' + sr].value == '001' and munkalap['P' + sr].value == 'HU' and munkalap['R' + sr].value == 'ZF49':
                    # JOHANS rövidkód
                    fuvarozokod = munkalap['I'+sr].value
                    if fuvarozokod == sapjohans:
                        if munkalap['B'+sr].value == gyvac:
                            utrange = utvacsheet[utdijbalfelso: utdijjobbalso]
                        if munkalap['B'+sr].value == gybere:
                            utrange = utberesheet[utdijbalfelso: utdijjobbalso]

                        for i in utrange:
                            if type(i[0].value) != NoneType:
                                szlevhelyseg = str(munkalap['O'+sr].value)
                                uthelyseg = str(i[1].value)
                                if szlevhelyseg.upper() == uthelyseg.upper():
                                    munkalap['AG'+sr].value = i[9].value
                                    break
                            else:
                                break

                        munkalap["AE" + str(
                            sr)].value = f'={munkalap["AH" + sr].value - munkalap["AG" + sr].value}'
                    # JOHANS rövidkód VÉGE

                    # Kemencepor
                    termek = munkalap['G'+sr].value
                    if termek == sapkemencepor:
                        if munkalap['B'+sr].value == gyvac:
                            utrange = utvacsheet[utdijbalfelso: utdijjobbalso]
                        if munkalap['B'+sr].value == gybere:
                            utrange = utberesheet[utdijbalfelso: utdijjobbalso]

                        for i in utrange:
                            if type(i[0].value) != NoneType:
                                szlevhelyseg = str(munkalap['O'+sr].value)
                                uthelyseg = str(i[1].value)
                                if szlevhelyseg.upper() == uthelyseg.upper():
                                    munkalap['AG'+sr].value = i[9].value * 2
                                    break
                            else:
                                break

                        munkalap["AE" + str(
                            sr)].value = f'={munkalap["AH" + sr].value - munkalap["AG" + sr].value}'
                    # Kemencepor VÉGE
                # ÖML - HU - ZF49 VÉGE

                # ÖML - SK - ZF49
                if munkalap['D' + sr].value == '001' and munkalap['P' + sr].value == 'SK' and munkalap['R' + sr].value == 'ZF49':
                    utrange = ut_skoml_sheet[skoml_utdijbalfelso: skoml_utdijjobbalso]

                    fuvarozokod = munkalap['I'+sr].value
                    arufogado = munkalap['M'+sr].value
                    afogado = int(arufogado)
                    for j in utrange:
                        # vizsgálja, hogy van-e adat a mezőben
                        if type(j[0].value) != NoneType:
                            if int(j[0].value) == afogado:
                                if fuvarozokod[0:4] == '1832':
                                    # oszlopok száma nullával kezdődik, így a táblában lévő
                                    # sorszámból egyet le kell vonni
                                    munkalap['AG'+sr].value = j[67].value
                                    break
                                else:
                                    munkalap['AG'+sr].value = j[66].value
                                    break
                        else:
                            break

                    munkalap["AE" + str(
                        sr)].value = f'={munkalap["AH" + sr].value - munkalap["AG" + sr].value}'
                # ÖML - SK - ZF49 - hosszúkód VÉGE

                # PAL - SK - ZF49
                if munkalap['D' + sr].value == '006' and munkalap['P' + sr].value == 'SK' and munkalap['R' + sr].value == 'ZF49':
                    utrange = ut_skpal_sheet[skpal_utdijbalfelso: skpal_utdijjobbalso]

                    fuvarozokod = munkalap['I'+sr].value
                    arufogado = munkalap['M'+sr].value
                    afogado = int(arufogado)
                    for j in utrange:
                        # vizsgálja, hogy van-e adat a mezőben
                        if type(j[0].value) != NoneType:
                            if int(j[0].value) == afogado:
                                if fuvarozokod == sapspeedline:
                                    munkalap['AG'+sr].value = 0
                                    break
                                if fuvarozokod == sapnordsped:
                                    munkalap['AG'+sr].value = 0
                                    break
                                if fuvarozokod == sappetranyi:
                                    # oszlopok száma nullával kezdődik, így a táblában lévő
                                    # sorszámból egyet le kell vonni
                                    munkalap['AG' +
                                             sr].value = j[22].value
                                    break
                                else:
                                    munkalap['AG' +
                                             sr].value = j[21].value
                                    break
                        else:
                            break

                    munkalap["AE" + str(
                        sr)].value = f'={munkalap["AH" + sr].value - munkalap["AG" + sr].value}'
                # PAL - SK - ZF49 - hosszúkód VÉGE

            # EXW
            elif munkalap['E' + sr].value == 'EXW':
                munkalap["AE" + sr].value = 0  # nettó fuvardíj
                munkalap["AG" + sr].value = 0  # útdíj
                munkalap["AH" + sr].value = 0  # fuvar- + útdíj
            # Incoterms a fentiek közül egyiksem
            else:
                munkalap["AE" + sr].value = 0
                munkalap["AG" + sr].value = 0
                munkalap["AH" + sr].value = 0

        # Excel fájlok bezárása
        wbook.save(saveas_szlevfajl)
        wbook.close()
        excelnaplo.info('Excel fájl elmentve, bezárva (szállítólevél adatok).')

        helyseg_book.close()
        excelnaplo.info('Excel fájl elmentve, bezárva (helységnevek).')

        utdij_book.close()
        excelnaplo.info('Excel fájl elmentve, bezárva (útdijak).')

        skoml_book.close()
        excelnaplo.info('Excel fájl elmentve, bezárva (SK-ÖML ártábla).')

        skpal_book.close()
        excelnaplo.info('Excel fájl elmentve, bezárva (SK-PAL ártábla).')

        # **** Excel fájl átalakítása VÉGE

    except FileNotFoundError:
        excelnaplo.error("Az XLSX fájl nem található: %s", str(szlevfajl))
        excelnaplo.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Az XLSX Fájl nem található!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)
    except Exception as merror:
        excelnaplo.error(
            'Ismeretlen hiba típusa, leírás: %s: %s', str(type(merror)), str(merror))
        excelnaplo.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Ismeretlen hiba!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)
    
    excelnaplo.info('Exel forrásfájl feldolgozása befejeződött')


def fejlecnevek(rs):
    """
    Fejlécnevek átírása, az adatbázis mezőkhöz igazítva
    Készült: 2022.06.16

    Paraméter:
        rs - Excelmunkalap változónév

    Utolsó módosítás dátuma: 2022.06.16
    Verzió: 01
    """
    rs["A1"].value = "ErtSzerv"
    rs["B1"].value = "Gyar"
    rs["C1"].value = "SzlevSzama"
    rs["D1"].value = "Csomagolas"
    rs["E1"].value = "Incoterms"
    rs["F1"].value = "Tetel"
    rs["G1"].value = "AnyagKod"
    rs["H1"].value = "Rendszam"
    rs["I1"].value = "FuvarozoKod"
    rs["J1"].value = "FuvarozoNeve"
    rs["K1"].value = "MegrendeloKod"
    rs["L1"].value = "MegrendeloNeve"
    rs["M1"].value = "ArufogadoKod"
    rs["N1"].value = "ArufogadoNeve"
    rs["O1"].value = "Helyseg"
    rs["P1"].value = "Orszag"
    rs["Q1"].value = "VevoKorzet"
    rs["R1"].value = "KondicioFajta"
    rs["S1"].value = "RendelesSzama"
    rs["T1"].value = "SzamlaSzama"
    rs["U1"].value = "SzlevDatum"
    rs["V1"].value = "Tonna"
    rs["W1"].value = "TonnaMertEgys"
    rs["X1"].value = "Tavolsag"
    rs["Y1"].value = "TavolsagMertEgys"
    rs["Z1"].value = "SzlaNettoErtek"
    rs["AA1"].value = "SzlaPenznem"
    rs["AB1"].value = "ATKm"
    rs["AC1"].value = "FuvarEgysAr"
    rs["AD1"].value = "FuvarPenznem"
    rs["AE1"].value = "FuvardijNetto"
    rs["AF1"].value = "UtdijKondicio"
    rs["AG1"].value = "Utdij"
    rs["AH1"].value = "FuvarUtdijBrutto"
    rs["AI1"].value = "EURArfolyam"
    rs["AJ1"].value = "RendelesTipus"
    rs["AK1"].value = "NettoFuvardij"
