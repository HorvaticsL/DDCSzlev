"""
Excel fájl tartalmának feldolgozása
A fájlban lévő adatokat úgy kell átalakítani, hogy
az adatbázisba lementhetők legyenek - helységnevek, fuvar-, útdíjak, stb.
Készült: 2022.06.13

Utolsó módosítás dátuma: 2022.07.06
verzió: 01

"""
#import openpyxl
from imp import load_dynamic
import sys
import ctypes
from types import NoneType
from openpyxl import Workbook, load_workbook
#from openpyxl.utils import get_column_letter

import ini_fajl as inif
import naplozas
import datumido as di


def excelfajl_modositas(initomb):
    # Utolsó módosítás dátuma: 2022.07.06
    # Naplózás beállítása
    excelnaplo = naplozas.naplolog()
    excelnaplo.info('Exel forrásfájl feldolgozása elindult')

    # INI fájladatok beolvasása
    prgneve = inif.initomb_eleme(initomb, 0)
    excelfile = inif.initomb_eleme(initomb, 3)
    excelmappa = inif.initomb_eleme(initomb, 4)
    wbsheetneve = inif.initomb_eleme(initomb, 5)

    helysegmappa = inif.initomb_eleme(initomb, 6)
    helysegfile = inif.initomb_eleme(initomb, 7)
    helysegsheet = inif.initomb_eleme(initomb, 8)
    helysegrange = inif.initomb_eleme(initomb, 9)

    szlevfajl = str(excelmappa) + str(excelfile)
    # print(szlevfajl)
    excelnaplo.info('Excelfájl (szállítólevelek): %s', str(excelfile))

    helysegfajl = str(helysegmappa) + str(helysegfile)
    # print(helysegfajl)
    excelnaplo.info('Excelfájl (helységnevek): %s', str(helysegfajl))

    # Fájl nevének módosítása (szállítólevelek)
    # Fájl végére beillesztésre kerül a napi dátum
    fileneve = excelfile[0:len(excelfile)-5]
    saveasexcelfile = fileneve + "_" + di.mainap("") + ".xlsx"
    excelnaplo.info("Excel fájl (SAVE AS) neve: %s", saveasexcelfile)
    saveas_szlevfajl = str(excelmappa) + str(saveasexcelfile)

    try:
        excelnaplo.info('Excel fájl megnyitása: szállítólevél adatok')
        # adatokat tartalmaézó Excel fájl
        wbook = load_workbook(filename=szlevfajl)
        munkalap = wbook[wbsheetneve]
        # print(munkalap['A1'].value)

        excelnaplo.info('Excel fájl megnyitása: helységnevek tábla')
        helyseg_book = load_workbook(filename=helysegfajl)
        helyseg_munkalap = helyseg_book[helysegsheet]
        hrange = helyseg_munkalap['G4':'H500']
        #print(hrange)

        maxsor = munkalap.max_row
        excelnaplo.info('Sorok száma (szállítólevelek): %s', str(maxsor))
        maxoszlop = munkalap.max_column
        excelnaplo.info('Oszlopok száma (szállítólevelek): %s', str(maxoszlop))

        #print("sorok: ", str(maxsor))
        #print("oszlopok: ", str(maxoszlop))

        # ***** Excel tábla fejléc

        # Fejléc beállítása az adatbázis mezőnekvekhez igazítva
        fejlecnevek(munkalap)
        excelnaplo.info("Fejlec megnevezés átírása")

        # **** Excel tábla fejléc VÉGE

        # **** Excel fájl átalakítása

        # Ciklus a soronkénti végig olvasáshoz, íráshoz
        for sor in range(2, maxsor):
            sr = str(sor)
            '''
            # Incoterms vizsgálata
            # CPT
            if munkalap['E' + sr].value == 'CPT':
                #ÖML - HU - ZF47
                if munkalap['D' + sr].value == '001' and munkalap['P' + sr].value == 'HU' and munkalap['R' + sr].value == 'ZF47':
                    munkalap["AE" + str(
                        sr)].value = f'={munkalap["AH" + sr].value - munkalap["AG" + sr].value}'

                #PAL - HU - ZF47
                if munkalap['D' + sr].value == '006' and munkalap['P' + sr].value == 'HU' and munkalap['R' + sr].value == 'ZF47':
                    munkalap["AE" + str(
                        sr)].value = f'={munkalap["AH" + sr].value - munkalap["AG" + sr].value}'

            # EXW
            elif munkalap['E' + sr].value == 'EXW':
                munkalap["AE" + sr].value = 0  # nettó fuvardíj
                munkalap["AG" + sr].value = 0  # útdíj
                munkalap["AH" + sr].value = 0  # fuvar- + útdíj
            # Incoterms a fentiek közül egyiksem
            else:
                munkalap["AE" + sr].value = 0
                munkalap["AG" + sr].value = 0
                munkalap["AH" + sr].value = 0
            '''
            # Helységnév átalakítása
            arufogado = munkalap['M'+sr].value
            afogado = int(arufogado)
            for j in hrange:
                # vizsgálja, hogy van-e adat a mezőben
                if type(j[0].value) != NoneType:
                    if int(j[0].value) == afogado:
                        munkalap['O'+sr].value = j[1].value
                        break
                else:
                    break

            # Helységnév átalakítása VÉGE

        wbook.save(saveas_szlevfajl)
        excelnaplo.info('Excel fájl elmentve')

        # **** Excel fájl átalakítása VÉGE

    except FileNotFoundError:
        excelnaplo.error("Az XLSX fájl nem található: %s", str(szlevfajl))
        excelnaplo.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Az XLSX Fájl nem található!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)
    except Exception as merror:
        excelnaplo.error(
            'Ismeretlen hiba típusa, leírás: %s: %s', str(type(merror)), str(merror))
        excelnaplo.warning("A program leállt!")
        MessageBox = ctypes.windll.user32.MessageBoxW
        MessageBox(
            None,
            "Ismeretlen hiba!\n\nRészletek a naplófájlban!",
            prgneve,
            0,
        )
        sys.exit(0)

    # CPT - VAN útdíj, NINCS útdíj
    # VAN útdíj = képelettel kiszámolni az nettó fuvardíjat
    # NINCS útdíj - Útdíj megkeresése, képlettel kiszámítani a nettó fuvardíjat

    excelnaplo.info('Exel forrásfájl feldolgozása befejeződött')


def fejlecnevek(rs):
    """
    Fejlécnevek átírása, az adatbázis mezőkhöz igazítva
    Készült: 2022.06.16

    Paraméter:
        rs - Excelmunkalap változónév

    Utolsó módosítás dátuma: 2022.06.16
    Verzió: 01
    """
    rs["A1"].value = "ErtSzerv"
    rs["B1"].value = "Gyar"
    rs["C1"].value = "SzlevSzama"
    rs["D1"].value = "Csomagolas"
    rs["E1"].value = "Incoterms"
    rs["F1"].value = "Tetel"
    rs["G1"].value = "AnyagKod"
    rs["H1"].value = "Rendszam"
    rs["I1"].value = "FuvarozoKod"
    rs["J1"].value = "FuvarozoNeve"
    rs["K1"].value = "MegrendeloKod"
    rs["L1"].value = "MegrendeloNeve"
    rs["M1"].value = "ArufogadoKod"
    rs["N1"].value = "ArufogadoNeve"
    rs["O1"].value = "Helyseg"
    rs["P1"].value = "Orszag"
    rs["Q1"].value = "VevoKorzet"
    rs["R1"].value = "KondicioFajta"
    rs["S1"].value = "RendelesSzama"
    rs["T1"].value = "SzamlaSzama"
    rs["U1"].value = "SzlevDatum"
    rs["V1"].value = "Tonna"
    rs["W1"].value = "TonnaMertEgys"
    rs["X1"].value = "Tavolsag"
    rs["Y1"].value = "TavolsagMertEgys"
    rs["Z1"].value = "SzlaNettoErtek"
    rs["AA1"].value = "SzlaPenznem"
    rs["AB1"].value = "ATKm"
    rs["AC1"].value = "FuvarEgysAr"
    rs["AD1"].value = "FuvarPenznem"
    rs["AE1"].value = "FuvardijNetto"
    rs["AF1"].value = "UtdijKondicio"
    rs["AG1"].value = "Utdij"
    rs["AH1"].value = "FuvarUtdijBrutto"
    rs["AI1"].value = "EURArfolyam"
    rs["AJ1"].value = "RendelesTipus"
    rs["AK1"].value = "NettoFuvardij"
